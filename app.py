import streamlit as st
import pandas as pd
import pdfplumber
import re
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== FUNGSI PARSING BNI ====================

def clean_bni_amount(raw):
    if not raw: return ""
    raw = str(raw).strip()
    if len(raw) >= 2 and len(raw) % 2 == 0:
        cand = "".join(raw[i] for i in range(0, len(raw), 2))
        if "".join(c+c for c in cand) == raw: return cand
    if '..' in raw:
        parts = raw.split('..')
        int_part = parts[0].replace(',,', ',')
        dec_part = parts[1] if len(parts) > 1 else '00'
        if len(dec_part) >= 2 and len(dec_part) % 2 == 0:
            dec_cand = "".join(dec_part[i] for i in range(0, len(dec_part), 2))
            if "".join(c+c for c in dec_cand) == dec_part: dec_part = dec_cand
        int_cleaned = int_part.replace(',', '')
        try:
            return f"{float(f'{int_cleaned}.{dec_part}'):,.2f}"
        except:
            return f"{int_cleaned}.{dec_part}"
    if ',,' in raw: raw = raw.replace(',,', ',')
    return raw

def clean_bni_text(s):
    if not s: return ""
    s = str(s).strip()
    if len(s) >= 2 and len(s) % 2 == 0:
        cand = "".join(s[i] for i in range(0, len(s), 2))
        if "".join(c+c for c in cand) == s: return cand
    return s

def to_float(s):
    if not s: return 0.0
    s = str(s).replace(',', '')
    try: return float(s)
    except:
        s = re.sub(r'[^\d.\-]', '', s)
        try: return float(s) if s else 0.0
        except: return 0.0

def audit_amount(raw_amt, prev_bal, curr_bal, dk):
    if prev_bal is None or curr_bal is None:
        return clean_bni_amount(raw_amt)
    target = abs(curr_bal - prev_bal)
    cleaned = clean_bni_amount(raw_amt)
    cleaned_val = to_float(cleaned)
    if abs(cleaned_val - target) < 1.0:
        return f"{cleaned_val:,.2f}"
    if target > 0:
        return f"{target:,.2f}"
    return cleaned

def extract_bni(pdf_file):
    all_data = []
    account_info = {}
    with pdfplumber.open(pdf_file) as pdf:
        c_acc, c_owner = "", ""
        running_bal = None
        current_ledger_bal = ""

        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                m = re.search(r'Account No\.?.*?:?\s*(\d{10})', text, re.I)
                if m:
                    new_acc = m.group(1)
                    if new_acc != c_acc:
                        c_acc = new_acc
                        running_bal = None

                    mn = re.search(rf'{c_acc}\s*/\s*(.*?)(?:\(|$)', text)
                    if mn: c_owner = mn.group(1).strip()

                if not account_info and c_acc:
                    m_comp = re.search(r'^\s*(PT\s+.+?)\s+Account', text, re.M)
                    company = m_comp.group(1).strip() if m_comp else ""
                    m_addr = re.search(r'^\s*(JL\s+.+?)\s+Account\s*Type', text, re.M | re.I)
                    address = m_addr.group(1).strip() if m_addr else ""
                    m_type = re.search(r'Account Type\s*:?\s*(\w+)', text, re.I)
                    acc_type = m_type.group(1) if m_type else ""
                    m_period = re.search(r'Period\s*:?\s*(.+?)$', text, re.M | re.I)
                    period = m_period.group(1).strip() if m_period else ""
                    m_ccy = re.search(r'\((USD|IDR)\)', text)
                    currency = m_ccy.group(1) if m_ccy else ""
                    account_info = {
                        "company": company, "address": address,
                        "acc_no": c_acc, "acc_name": c_owner,
                        "acc_type": acc_type, "period": period, "currency": currency
                    }

            table = page.extract_table()
            if not table: continue

            col, h_idx = {}, -1
            num_cols = len(table[0]) if table else 0
            for i, row in enumerate(table):
                row_str = " ".join([clean_bni_text(str(c)).upper() for c in row if c])
                if "POSTING DATE" in row_str:
                    h_idx = i
                    for j, cell in enumerate(row):
                        t = clean_bni_text(str(cell)).upper()
                        if "POSTING" in t: col["post"] = j
                        elif "EFFECTIVE" in t: col["eff"] = j
                        elif "BRANCH" in t: col["branch"] = j
                        elif "JOURNAL" in t: col["journal"] = j
                        elif "DESCRIPTION" in t: col["desc"] = j
                        elif "AMOUNT" in t: col["amt"] = j
                        elif "DB/CR" in t: col["dk"] = j
                        elif "BALANCE" in t: col["bal"] = j
                    break

            if h_idx == -1 and not col:
                if num_cols == 8:
                    col = {"post": 0, "eff": 1, "branch": 2, "journal": 3,
                           "desc": 4, "amt": 5, "dk": 6, "bal": 7}
                elif num_cols == 10:
                    col = {"post": 1, "eff": 2, "branch": 3, "journal": 4,
                           "desc": 5, "amt": 6, "dk": 7, "bal": 9}
                else: continue

            active_col = col
            if not active_col: continue
            start = h_idx + 1 if h_idx != -1 else 0

            for row in table[start:]:
                def val(key):
                    idx = active_col.get(key)
                    if idx is not None and idx < len(row) and row[idx]:
                        return str(row[idx]).strip()
                    return ""

                row_text = " ".join([str(x) for x in row if x])

                if "Ledger Balance" in row_text:
                    for cell in reversed(row):
                        if cell and re.search(r'[\d,.]+', str(cell)) and '/' not in str(cell):
                            running_bal = to_float(str(cell).replace(',', ''))
                            current_ledger_bal = f"{running_bal:,.2f}"
                            if "starting_balance" not in account_info:
                                account_info["starting_balance"] = current_ledger_bal
                            # Sisipkan baris "Saldo Awal" ke data
                            all_data.append({
                                "Posting Date": "", "Effective Date": "",
                                "Branch": "", "Journal": "",
                                "Transaction Description": "Saldo Awal",
                                "Amount": "", "DB/CR": "",
                                "Balance": current_ledger_bal
                            })
                            break
                    continue

                if "Ending Balance" in row_text: continue

                p_date = ""
                for cell in row[:3]:
                    c = str(cell).strip() if cell else ""
                    if c and re.search(r'\d{2}/\d{2}/\d{4}', c):
                        p_date = c
                        break

                if p_date:
                    raw_amt, raw_bal = val("amt"), val("bal")
                    dk = clean_bni_text(val("dk"))
                    curr_bal = to_float(raw_bal.replace(',', '')) if raw_bal else None
                    final_amt = audit_amount(raw_amt, running_bal, curr_bal, dk)
                    eff = val("eff")
                    if not eff or not re.search(r'\d{2}/\d{2}/\d{4}', eff): eff = p_date

                    all_data.append({
                        "Posting Date": p_date, "Effective Date": eff,
                        "Branch": clean_bni_text(val("branch")).replace('\n', ' '),
                        "Journal": clean_bni_text(val("journal")),
                        "Transaction Description": val("desc").replace('\n', ' '),
                        "Amount": final_amt, "DB/CR": dk,
                        "Balance": f"{curr_bal:,.2f}" if curr_bal else ""
                    })
                    if curr_bal: running_bal = curr_bal

                elif all_data and any(row):
                    last = all_data[-1]
                    b, d = val("branch"), val("desc")
                    if b and not re.search(r'(ENDING|TOTAL|LEDGER)', b, re.I):
                        last["Branch"] = (last["Branch"] + " " + clean_bni_text(b).replace('\n', ' ')).strip()
                    if d and not re.search(r'(ENDING|TOTAL|LEDGER)', d, re.I):
                        last["Transaction Description"] = (last["Transaction Description"] + " " + d.replace('\n', ' ')).strip()

    return all_data, account_info


def create_excel_bytes(df, info):
    """Buat file Excel dengan header info dan styling, return sebagai bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mutasi'

    title_font = Font(bold=True, size=14, color="FF6600")
    label_font = Font(bold=True, size=10)
    value_font = Font(size=10)

    ws.merge_cells('A1:I1')
    ws['A1'] = 'ACCOUNT STATEMENT'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    ws['A3'] = info.get('company', '')
    ws['A3'].font = Font(bold=True, size=10)
    ws['A4'] = info.get('address', '')
    ws['A4'].font = value_font

    ws['D3'] = 'Account No.'
    ws['D3'].font = label_font
    ws['E3'] = f": {info.get('acc_no', '')} / {info.get('acc_name', '')}"
    ws['E3'].font = value_font

    ws['D4'] = 'Account Type'
    ws['D4'].font = label_font
    ws['E4'] = f": {info.get('acc_type', '')}   ({info.get('currency', '')})"
    ws['E4'].font = value_font

    ws['D5'] = 'Period'
    ws['D5'].font = label_font
    ws['E5'] = f": {info.get('period', '')}"
    ws['E5'].font = value_font

    ws['D6'] = 'Saldo Awal'
    ws['D6'].font = label_font
    ws['E6'] = f": {info.get('starting_balance', '-')}"
    ws['E6'].font = value_font

    start_row = 9
    header_fill = PatternFill(start_color='FF6600', end_color='FF6600', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for ci, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=ci, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    for ri, row_data in enumerate(df.values, start_row + 1):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.font = value_font

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = len(str(col_name))
        for val in df[col_name].astype(str):
            if len(val) > max_len: max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    ws.freeze_panes = f'A{start_row + 1}'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ==================== STREAMLIT APP ====================

st.set_page_config(
    page_title="Koranizer - BNI PDF Converter",
    page_icon="🏦",
    layout="wide"
)

# Custom CSS
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

    .stApp {
        font-family: 'Inter', sans-serif;
    }

    .main-header {
        background: linear-gradient(135deg, #FF6600 0%, #FF8533 50%, #FFa366 100%);
        padding: 2rem 2.5rem;
        border-radius: 16px;
        margin-bottom: 2rem;
        box-shadow: 0 8px 32px rgba(255, 102, 0, 0.2);
    }
    .main-header h1 {
        color: white;
        font-size: 2.2rem;
        font-weight: 700;
        margin: 0;
    }
    .main-header p {
        color: rgba(255,255,255,0.9);
        font-size: 1.05rem;
        margin: 0.3rem 0 0 0;
    }

    .info-card {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        color: white;
        padding: 1.5rem;
        border-radius: 12px;
        margin-bottom: 1.5rem;
        border-left: 4px solid #FF6600;
    }
    .info-card .label { color: #FFa366; font-weight: 600; font-size: 0.85rem; }
    .info-card .value { color: white; font-size: 1.05rem; margin-bottom: 0.4rem; }

    .stat-box {
        background: linear-gradient(135deg, #0f3460 0%, #16213e 100%);
        padding: 1.2rem;
        border-radius: 12px;
        text-align: center;
        border: 1px solid rgba(255, 102, 0, 0.3);
    }
    .stat-box .number { color: #FF6600; font-size: 1.8rem; font-weight: 700; }
    .stat-box .label { color: #ccc; font-size: 0.8rem; margin-top: 0.2rem; }

    div[data-testid="stFileUploader"] {
        border: 2px dashed #FF6600;
        border-radius: 12px;
        padding: 1rem;
    }

    .stDataFrame { border-radius: 8px; overflow: hidden; }
</style>
""", unsafe_allow_html=True)

# Header
st.markdown("""
<div class="main-header">
    <h1>🏦 Koranizer</h1>
    <p>Konversi Rekening Koran BNI PDF ke Excel — Akurat & Otomatis</p>
</div>
""", unsafe_allow_html=True)

# Upload
uploaded_files = st.file_uploader(
    "📄 Upload file PDF Rekening Koran BNI",
    type=["pdf"],
    accept_multiple_files=True,
    help="Anda bisa upload satu atau lebih file PDF BNI sekaligus."
)

if uploaded_files:
    for uploaded_file in uploaded_files:
        st.markdown(f"### 📋 {uploaded_file.name}")

        with st.spinner(f"⏳ Memproses {uploaded_file.name}..."):
            data, info = extract_bni(uploaded_file)

        if not data:
            st.error(f"❌ Tidak ada transaksi yang ditemukan di {uploaded_file.name}")
            continue

        df = pd.DataFrame(data)
        df["Branch"] = df["Branch"].str.replace(r'\s+', ' ', regex=True).str.strip()
        df["Transaction Description"] = df["Transaction Description"].str.replace(r'\s+', ' ', regex=True).str.strip()
        df = df.fillna("")

        # Info Akun
        st.markdown(f"""
        <div class="info-card">
            <div class="label">PERUSAHAAN</div>
            <div class="value">{info.get('company', '-')}</div>
            <div class="label">ALAMAT</div>
            <div class="value">{info.get('address', '-')}</div>
            <div class="label">NOMOR REKENING</div>
            <div class="value">{info.get('acc_no', '-')} / {info.get('acc_name', '-')}</div>
            <div class="label">TIPE & MATA UANG</div>
            <div class="value">{info.get('acc_type', '-')} ({info.get('currency', '-')})</div>
            <div class="label">PERIODE</div>
            <div class="value">{info.get('period', '-')}</div>
            <div class="label">SALDO AWAL</div>
            <div class="value">{info.get('starting_balance', '-')}</div>
        </div>
        """, unsafe_allow_html=True)

        # Statistik
        total_trx = len(df)
        total_debit = len(df[df["DB/CR"] == "D"])
        total_kredit = len(df[df["DB/CR"] == "K"])

        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown(f'<div class="stat-box"><div class="number">{total_trx}</div><div class="label">Total Transaksi</div></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="stat-box"><div class="number">{total_debit}</div><div class="label">Debit (D)</div></div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div class="stat-box"><div class="number">{total_kredit}</div><div class="label">Kredit (K)</div></div>', unsafe_allow_html=True)

        st.markdown("---")

        # Preview Data
        st.markdown("#### 📊 Preview Data")
        st.dataframe(df, use_container_width=True, height=400)

        # Download Excel
        excel_bytes = create_excel_bytes(df, info)
        out_name = uploaded_file.name.replace(".pdf", "_RAPI.xlsx")

        st.download_button(
            label=f"📥 Download Excel: {out_name}",
            data=excel_bytes,
            file_name=out_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )

        st.markdown("---")

else:
    st.markdown("""
    <div style="text-align:center; padding: 3rem; color: #888;">
        <p style="font-size: 3rem;">📄</p>
        <p style="font-size: 1.1rem;">Upload file PDF BNI di atas untuk memulai konversi</p>
        <p style="font-size: 0.9rem; color: #666;">Mendukung rekening USD dan IDR, multi-halaman, serta multi-akun dalam satu PDF</p>
    </div>
    """, unsafe_allow_html=True)

# Footer
st.markdown("""
<div style="text-align:center; padding: 1.5rem; margin-top: 2rem; color: #555; font-size: 0.8rem;">
    Koranizer v2.0 — BNI PDF to Excel Converter | Built with Streamlit
</div>
""", unsafe_allow_html=True)
